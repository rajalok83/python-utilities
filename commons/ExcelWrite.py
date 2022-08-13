# Install  openpyxl
# pip install openpyxl (2.7) or pip3 install openpyxl (3+)

import logging
import sys
import traceback
from os import path

import openpyxl


class ExcelProcess:
    wb_path = None
    log_enabled = True

    def read_from_excel(self, sht_nm, get_typ, rng=None, row_num=None, col_num=None):
        if not (path.exists(self.wb_path)):
            self.log_me("File {} does not exists.".format(self.wb_path), True, 8)
        self.log_me("Beginning to read file {}".format(self.wb_path), False, -1)
        try:
            wb_obj = openpyxl.load_workbook(self.wb_path)
            ws = wb_obj[sht_nm]
            self.log_me("Sheet {} found in file {}".format(sht_nm, self.wb_path), False, -1)
            if get_typ == "all":
                # To get all records
                return ws.rows
            elif get_typ == "range":
                # To get Range of records like in excel
                return ws[rng]
            elif get_typ == "single":
                # To get one record
                return ws.cell(row=row_num, column=col_num)
            else:
                return None
        except KeyError:
            traceback.print_exc(file=sys.stdout)
            self.log_me("Sheet name {} not found in file {}".format(sht_nm, self.wb_path), True, 8)
        except TypeError:
            traceback.print_exc(file=sys.stdout)
            if get_typ == "range":
                self.log_me("Format of parameter rng must be like 'A1:B2' and is compulsary for get_typ range", True, 8)
            elif get_typ == "single":
                self.log_me("Parameter row_num and col_num must be of type int and compulsary for get_typ single", True,
                            8)
        return

    @staticmethod
    def write_to_excel():
        wb = openpyxl.Workbook()
        sheet = wb.active
        c1 = sheet.cell(row=1, column=1)
        c1.value = "ANKIT"
        c2 = sheet.cell(row=1, column=2)
        c2.value = "RAI"
        c3 = sheet['A2']
        c3.value = "RAHUL"
        c4 = sheet['B2']
        c4.value = "RAI"
        return

    @staticmethod
    def is_sheet_present(wb_obj, sht_nm):
        print(wb_obj.get_index(sht_nm))
        try:
            print(wb_obj.get_index(sht_nm))
            if wb_obj.get_index(sht_nm) >= 0:
                return True
            else:
                return False
        except ValueError:
            return False

    def log_me(self, msg, prnt, exit_now):
        if self.log_enabled:
            if exit_now > -1:
                logging.error(msg)
            else:
                logging.info(msg)
        if prnt:
            print(msg)
        if exit_now > -1:
            exit(exit_now)
        return

    def __init__(self, wb_path, log_enabled):
        print("Logging Enabled:{}".format(log_enabled))
        if log_enabled:
            logging.basicConfig(filename="d:\\Users\\hemu\\Desktop\\pylog.log", format='%(asctime)s %(message)s',
                                datefmt='%m/%d/%Y %I:%M:%S %p')
            self.log_me("Logging Enabled: {}".format(log_enabled), False, -1)
            self.log_me("Setting workboook path to {}".format(wb_path), False, -1)
        self.wb_path = wb_path
        self.log_enabled = log_enabled
        return


if __name__ == '__main__':
    xlPrc = ExcelProcess("D:\\Users\\hemu\\Desktop\\demo.xlsx", True)
    readRows = xlPrc.read_from_excel("Sheet", "single", "A1:B3")
    # For single cell
    # print(readRows.value)
    # For all and ranges
    for row in readRows:
        for col in row:
            print(col.value)
